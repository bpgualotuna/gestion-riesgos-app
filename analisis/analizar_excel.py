"""
Script para analizar en profundidad el archivo Excel de gestión de riesgo de Talento Humano
"""
import sys
import os
from pathlib import Path

try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("openpyxl no está instalado. Instalando...")
    os.system("pip install openpyxl")
    try:
        import openpyxl
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        HAS_OPENPYXL = True
    except:
        print("No se pudo instalar openpyxl. Usando alternativa...")
        HAS_OPENPYXL = False

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

def analizar_excel(archivo_path):
    """Analiza un archivo Excel en profundidad"""
    
    if not os.path.exists(archivo_path):
        print(f"ERROR: El archivo {archivo_path} no existe")
        return
    
    print("=" * 80)
    print(f"ANÁLISIS PROFUNDO DEL ARCHIVO EXCEL")
    print(f"Archivo: {archivo_path}")
    print("=" * 80)
    print()
    
    # Intentar con openpyxl (para .xlsx y .xlsm)
    if HAS_OPENPYXL:
        try:
            wb = load_workbook(archivo_path, data_only=False, keep_vba=True)
            analizar_con_openpyxl(wb, archivo_path)
            return
        except Exception as e:
            print(f"Error con openpyxl: {e}")
            print("Intentando con xlrd...")
    
    # Si falla, intentar con xlrd
    if HAS_XLRD:
        try:
            wb = xlrd.open_workbook(archivo_path)
            analizar_con_xlrd(wb, archivo_path)
            return
        except Exception as e:
            print(f"Error con xlrd: {e}")
    
    print("ERROR: No se pudo abrir el archivo con ninguna librería disponible")

def analizar_con_openpyxl(wb, archivo_path):
    """Análisis detallado usando openpyxl"""
    
    print(f"Tipo de archivo: {type(wb).__name__}")
    print(f"¿Tiene macros?: {wb.vba_archive is not None}")
    print()
    
    # Información general
    print("=" * 80)
    print("1. INFORMACIÓN GENERAL")
    print("=" * 80)
    print(f"Total de hojas: {len(wb.sheetnames)}")
    print(f"Nombres de hojas: {', '.join(wb.sheetnames)}")
    print()
    
    # Análisis de cada hoja
    for idx, sheet_name in enumerate(wb.sheetnames, 1):
        print("=" * 80)
        print(f"2.{idx}. ANÁLISIS DE LA HOJA: '{sheet_name}'")
        print("=" * 80)
        
        ws = wb[sheet_name]
        
        # Información básica de la hoja
        print(f"Dimensiones: {ws.dimensions}")
        print(f"Fila máxima: {ws.max_row}")
        print(f"Columna máxima: {ws.max_column}")
        print()
        
        # Buscar tablas estructuradas
        if hasattr(ws, 'tables'):
            print(f"Tablas estructuradas encontradas: {len(ws.tables)}")
            for table_name, table in ws.tables.items():
                print(f"  - Tabla: {table_name}")
                print(f"    Rango: {table.ref}")
                print(f"    Columnas: {len(table.tableColumns)}")
                for col in table.tableColumns:
                    print(f"      * {col.name}")
            print()
        
        # Análisis de fórmulas
        print("Fórmulas encontradas (primeras 20):")
        formulas_count = 0
        formulas_samples = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f' and cell.value:
                    formulas_count += 1
                    if len(formulas_samples) < 20:
                        formulas_samples.append({
                            'cell': f"{cell.coordinate}",
                            'formula': cell.value,
                            'value': cell.value if cell.data_type != 'f' else ws[cell.coordinate].value
                        })
        
        print(f"Total de fórmulas en la hoja: {formulas_count}")
        for sample in formulas_samples[:20]:
            print(f"  {sample['cell']}: {sample['formula']}")
        print()
        
        # Análisis de datos estructurados
        print("Estructura de datos (primeras 10 filas):")
        header_row = None
        data_start = None
        
        # Buscar fila de encabezados
        for row_idx in range(1, min(11, ws.max_row + 1)):
            row_values = [cell.value for cell in ws[row_idx]]
            if any(v is not None and str(v).strip() for v in row_values):
                if header_row is None:
                    header_row = row_idx
                    print(f"  Fila de encabezados (probable): {row_idx}")
                    print(f"  Encabezados: {[str(v)[:30] if v else '' for v in row_values[:15]]}")
                    data_start = row_idx + 1
                    break
        
        # Mostrar algunas filas de datos
        if data_start:
            print(f"\n  Datos (filas {data_start} a {min(data_start + 5, ws.max_row)}):")
            for row_idx in range(data_start, min(data_start + 5, ws.max_row + 1)):
                row_values = [str(cell.value)[:30] if cell.value else '' for cell in ws[row_idx]]
                non_empty = [v for v in row_values if v.strip()]
                if non_empty:
                    print(f"    Fila {row_idx}: {non_empty[:10]}")
        print()
        
        # Buscar validaciones de datos
        if hasattr(ws, 'data_validations'):
            validations = list(ws.data_validations.dataValidation)
            if validations:
                print(f"Validaciones de datos encontradas: {len(validations)}")
                for val in validations[:5]:
                    print(f"  - Rango: {val.ranges}, Tipo: {val.type}, Fórmula: {val.formula1}")
                print()
        
        # Buscar rangos con nombres
        if hasattr(wb, 'defined_names'):
            sheet_names = [name for name in wb.defined_names.localnames(ws.title)]
            if sheet_names:
                print(f"Nombres definidos en esta hoja: {len(sheet_names)}")
                for name in sheet_names[:10]:
                    try:
                        named_range = wb.defined_names[name]
                        print(f"  - {name}: {named_range.attr_text}")
                    except:
                        print(f"  - {name}: (no se pudo obtener)")
                print()
        
        # Análisis de formato condicional
        if hasattr(ws, 'conditional_formatting'):
            cf_count = len(ws.conditional_formatting._cf_rules)
            if cf_count > 0:
                print(f"Formatos condicionales encontrados: {cf_count}")
                for cf_rule in list(ws.conditional_formatting._cf_rules.items())[:5]:
                    print(f"  - Rango: {cf_rule[0]}, Reglas: {len(cf_rule[1])}")
                print()
        
        print()

def analizar_con_xlrd(wb, archivo_path):
    """Análisis usando xlrd (para archivos .xls antiguos)"""
    print(f"Total de hojas: {wb.nsheets}")
    for i in range(wb.nsheets):
        sheet = wb.sheet_by_index(i)
        print(f"\nHoja {i+1}: {sheet.name}")
        print(f"  Filas: {sheet.nrows}, Columnas: {sheet.ncols}")

def analizar_macros(archivo_path):
    """Intenta extraer información sobre macros VBA"""
    print("=" * 80)
    print("3. ANÁLISIS DE MACROS VBA")
    print("=" * 80)
    
    if HAS_OPENPYXL:
        try:
            wb = load_workbook(archivo_path, keep_vba=True)
            if wb.vba_archive:
                print("✓ El archivo contiene macros VBA")
                # Intentar listar módulos
                try:
                    # Esto requiere acceso al archivo zip interno
                    import zipfile
                    with zipfile.ZipFile(archivo_path, 'r') as zip_ref:
                        vba_files = [f for f in zip_ref.namelist() if 'vba' in f.lower() or f.endswith('.bas') or f.endswith('.cls')]
                        if vba_files:
                            print(f"Archivos VBA encontrados: {len(vba_files)}")
                            for vba_file in vba_files[:10]:
                                print(f"  - {vba_file}")
                except Exception as e:
                    print(f"  (No se pudieron listar los módulos: {e})")
            else:
                print("✗ El archivo NO contiene macros VBA")
        except Exception as e:
            print(f"Error al analizar macros: {e}")
    print()

if __name__ == "__main__":
    archivo = r"Herramienta de gestión de riesgo Talento Humano V1 revisada.xlsm"
    
    if not os.path.exists(archivo):
        print(f"ERROR: No se encuentra el archivo: {archivo}")
        sys.exit(1)
    
    analizar_excel(archivo)
    analizar_macros(archivo)
    
    print("=" * 80)
    print("ANÁLISIS COMPLETADO")
    print("=" * 80)

