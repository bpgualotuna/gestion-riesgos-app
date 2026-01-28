"""
Análisis detallado del archivo Excel de gestión de riesgo de Talento Humano
"""
import os
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import zipfile

def analizar_excel_detallado(archivo_path):
    """Análisis completo del archivo Excel"""
    
    print("=" * 100)
    print("ANÁLISIS PROFUNDO: Herramienta de gestión de riesgo Talento Humano")
    print("=" * 100)
    print()
    
    wb = load_workbook(archivo_path, data_only=False, keep_vba=True)
    
    # Información general
    print("INFORMACIÓN GENERAL")
    print("-" * 100)
    print(f"Total de hojas: {len(wb.sheetnames)}")
    print(f"Contiene macros VBA: {'Sí' if wb.vba_archive else 'No'}")
    print()
    
    # Análisis detallado de cada hoja
    analisis_hojas = {}
    
    for sheet_name in wb.sheetnames:
        print("=" * 100)
        print(f"HOJA: {sheet_name}")
        print("=" * 100)
        
        ws = wb[sheet_name]
        info_hoja = {
            'nombre': sheet_name,
            'max_row': ws.max_row,
            'max_column': ws.max_column,
            'dimensiones': ws.dimensions,
            'formulas': [],
            'estructura_datos': {},
            'validaciones': [],
            'formatos_condicionales': [],
            'tablas': []
        }
        
        # Tablas estructuradas
        if hasattr(ws, 'tables') and ws.tables:
            for table_name, table in ws.tables.items():
                info_tabla = {
                    'nombre': table_name,
                    'rango': table.ref,
                    'columnas': [col.name for col in table.tableColumns]
                }
                info_hoja['tablas'].append(info_tabla)
                print(f"  Tabla estructurada: {table_name}")
                print(f"    Rango: {table.ref}")
                print(f"    Columnas: {len(table.tableColumns)}")
        
        # Fórmulas importantes
        formulas_importantes = []
        formulas_count = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f' and cell.value:
                    formulas_count += 1
                    formula = str(cell.value)
                    # Filtrar fórmulas importantes (que no sean solo referencias simples)
                    if any(op in formula for op in ['SUM', 'IF', 'VLOOKUP', 'INDEX', 'MATCH', 'COUNT', 'AVERAGE', 'MAX', 'MIN']):
                        if len(formulas_importantes) < 50:
                            formulas_importantes.append({
                                'celda': cell.coordinate,
                                'formula': formula[:200]
                            })
        
        info_hoja['total_formulas'] = formulas_count
        info_hoja['formulas'] = formulas_importantes[:30]  # Limitar a 30
        
        if formulas_count > 0:
            print(f"  Total de fórmulas: {formulas_count}")
            print(f"  Fórmulas importantes encontradas: {len(formulas_importantes)}")
            for f in formulas_importantes[:10]:
                print(f"    {f['celda']}: {f['formula'][:80]}")
        
        # Estructura de datos - buscar encabezados y datos
        print(f"\n  Estructura de datos:")
        print(f"    Dimensiones: {ws.dimensions}")
        print(f"    Filas: {ws.max_row}, Columnas: {ws.max_column}")
        
        # Buscar fila de encabezados
        header_found = False
        for row_idx in range(1, min(20, ws.max_row + 1)):
            row_values = []
            for col_idx in range(1, min(ws.max_column + 1, 20)):
                cell = ws.cell(row=row_idx, column=col_idx)
                row_values.append(str(cell.value) if cell.value else '')
            
            non_empty = [v for v in row_values if v.strip()]
            if len(non_empty) >= 3:  # Si hay al menos 3 celdas con datos
                if not header_found:
                    print(f"    Posible fila de encabezados (fila {row_idx}):")
                    for i, val in enumerate(non_empty[:15], 1):
                        print(f"      Col {i}: {val[:50]}")
                    header_found = True
                    info_hoja['estructura_datos']['fila_encabezados'] = row_idx
                    info_hoja['estructura_datos']['encabezados'] = non_empty[:20]
                    break
        
        # Muestra de datos
        if header_found:
            data_rows = []
            for row_idx in range(info_hoja['estructura_datos']['fila_encabezados'] + 1, 
                               min(info_hoja['estructura_datos']['fila_encabezados'] + 6, ws.max_row + 1)):
                row_data = []
                for col_idx in range(1, min(ws.max_column + 1, 15)):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    val = str(cell.value)[:30] if cell.value else ''
                    if val.strip():
                        row_data.append(val)
                if row_data:
                    data_rows.append({'fila': row_idx, 'datos': row_data[:10]})
            
            if data_rows:
                print(f"\n    Muestra de datos (primeras {len(data_rows)} filas):")
                for dr in data_rows[:3]:
                    print(f"      Fila {dr['fila']}: {dr['datos']}")
        
        # Validaciones de datos
        if hasattr(ws, 'data_validations'):
            validations = list(ws.data_validations.dataValidation)
            if validations:
                print(f"\n  Validaciones de datos: {len(validations)}")
                for val in validations[:5]:
                    print(f"    Rango: {val.ranges}, Tipo: {val.type}")
                    info_hoja['validaciones'].append({
                        'rango': str(val.ranges),
                        'tipo': val.type
                    })
        
        # Formatos condicionales
        if hasattr(ws, 'conditional_formatting'):
            cf_rules = ws.conditional_formatting._cf_rules
            if cf_rules:
                print(f"\n  Formatos condicionales: {len(cf_rules)}")
                for range_str, rules in list(cf_rules.items())[:5]:
                    print(f"    Rango: {range_str}, Reglas: {len(rules)}")
                    info_hoja['formatos_condicionales'].append({
                        'rango': range_str,
                        'num_reglas': len(rules)
                    })
        
        analisis_hojas[sheet_name] = info_hoja
        print()
    
    # Análisis de macros
    print("=" * 100)
    print("ANÁLISIS DE MACROS VBA")
    print("=" * 100)
    
    if wb.vba_archive:
        try:
            with zipfile.ZipFile(archivo_path, 'r') as zip_ref:
                vba_files = [f for f in zip_ref.namelist() 
                           if any(x in f.lower() for x in ['vba', '.bas', '.cls', '.frm', 'xl/vba'])]
                if vba_files:
                    print(f"Archivos VBA encontrados: {len(vba_files)}")
                    for vba_file in vba_files[:20]:
                        print(f"  - {vba_file}")
        except Exception as e:
            print(f"Error al listar macros: {e}")
    else:
        print("No se encontraron macros VBA")
    
    print()
    
    # Resumen ejecutivo
    print("=" * 100)
    print("RESUMEN EJECUTIVO PARA MIGRACIÓN A WEB")
    print("=" * 100)
    print()
    print("COMPONENTES IDENTIFICADOS:")
    print()
    print("1. ESTRUCTURA DE DATOS:")
    print(f"   - {len(wb.sheetnames)} hojas de cálculo")
    print(f"   - Hojas principales identificadas:")
    
    hojas_principales = {
        'Ficha': 'Datos principales del proceso',
        'Inventario de Normatividad': 'Catálogo de normativas',
        'Análisis de Proceso': 'Análisis del proceso',
        'Identificación': 'Identificación de riesgos',
        'Evaluación': 'Evaluación de riesgos',
        'Mapa de riesgos': 'Visualización de riesgos',
        'Priorización y Respuesta': 'Gestión de respuestas',
        'Formulas': 'Fórmulas de cálculo',
        'Parámetros de Valoración': 'Parámetros del sistema',
        'Listas': 'Listas desplegables',
        'Tabla de atribuciones': 'Atribuciones',
        'Tipologias': 'Clasificaciones de riesgos'
    }
    
    for hoja in wb.sheetnames:
        desc = hojas_principales.get(hoja.split('.')[-1].strip(), 'Datos auxiliares')
        formulas = analisis_hojas[hoja]['total_formulas']
        tablas = len(analisis_hojas[hoja]['tablas'])
        print(f"     • {hoja}: {desc}")
        if formulas > 0:
            print(f"       - {formulas} fórmulas")
        if tablas > 0:
            print(f"       - {tablas} tablas estructuradas")
    
    print()
    print("2. FUNCIONALIDADES:")
    total_formulas = sum(h['total_formulas'] for h in analisis_hojas.values())
    total_validaciones = sum(len(h['validaciones']) for h in analisis_hojas.values())
    total_tablas = sum(len(h['tablas']) for h in analisis_hojas.values())
    
    print(f"   - {total_formulas} fórmulas en total")
    print(f"   - {total_validaciones} validaciones de datos")
    print(f"   - {total_tablas} tablas estructuradas")
    print(f"   - Macros VBA: {'Sí' if wb.vba_archive else 'No'}")
    print()
    
    print("3. RECOMENDACIONES PARA MIGRACIÓN:")
    print("   - Convertir tablas estructuradas a tablas de base de datos")
    print("   - Implementar fórmulas como funciones del backend o frontend")
    print("   - Reemplazar validaciones de Excel por validaciones en formularios web")
    print("   - Convertir macros VBA a scripts del lado del servidor o cliente")
    print("   - Implementar formato condicional como estilos CSS dinámicos")
    print("   - Crear API REST para operaciones CRUD")
    print("   - Implementar sistema de permisos y roles")
    print()
    
    # Guardar análisis en JSON
    with open('analisis_excel.json', 'w', encoding='utf-8') as f:
        json.dump({
            'total_hojas': len(wb.sheetnames),
            'hojas': analisis_hojas,
            'tiene_macros': wb.vba_archive is not None
        }, f, indent=2, ensure_ascii=False)
    
    print("Análisis guardado en: analisis_excel.json")
    print()

if __name__ == "__main__":
    archivo = r"Herramienta de gestión de riesgo Talento Humano V1 revisada.xlsm"
    analizar_excel_detallado(archivo)

