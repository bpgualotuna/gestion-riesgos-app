"""
Análisis completo hoja por hoja del archivo Excel
"""
import os
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from collections import defaultdict

def analizar_hoja_completa(ws, nombre_hoja):
    """Análisis exhaustivo de una hoja"""
    
    analisis = {
        'nombre': nombre_hoja,
        'dimensiones': ws.dimensions,
        'max_row': ws.max_row,
        'max_column': ws.max_column,
        'estructura': {},
        'formulas': [],
        'validaciones': [],
        'formatos_condicionales': [],
        'referencias_cruzadas': [],
        'datos_muestra': [],
        'encabezados': [],
        'tablas': []
    }
    
    print("\n" + "="*120)
    print(f"HOJA: {nombre_hoja}")
    print("="*120)
    print(f"Dimensiones: {ws.dimensions}")
    print(f"Filas: {ws.max_row}, Columnas: {ws.max_column}")
    print()
    
    # 1. TABLAS ESTRUCTURADAS
    if hasattr(ws, 'tables') and ws.tables:
        print("TABLAS ESTRUCTURADAS:")
        for table_name, table in ws.tables.items():
            tabla_info = {
                'nombre': table_name,
                'rango': table.ref,
                'columnas': []
            }
            for col in table.tableColumns:
                col_info = {
                    'nombre': col.name,
                    'id': col.id
                }
                tabla_info['columnas'].append(col_info)
                print(f"  - {table_name}: {table.ref}")
                print(f"    Columnas: {len(table.tableColumns)}")
            analisis['tablas'].append(tabla_info)
        print()
    
    # 2. BUSCAR ENCABEZADOS Y ESTRUCTURA
    print("ESTRUCTURA DE DATOS:")
    header_row = None
    for row_idx in range(1, min(30, ws.max_row + 1)):
        row_values = []
        for col_idx in range(1, min(ws.max_column + 1, 30)):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = str(cell.value) if cell.value is not None else ''
            row_values.append(val)
        
        non_empty = [v for v in row_values if v and v.strip()]
        if len(non_empty) >= 3 and header_row is None:
            header_row = row_idx
            analisis['estructura']['fila_encabezados'] = row_idx
            analisis['encabezados'] = row_values[:ws.max_column]
            print(f"  Fila de encabezados: {row_idx}")
            print(f"  Encabezados encontrados ({len(non_empty)}):")
            for i, val in enumerate(non_empty[:20], 1):
                print(f"    {i}. {val[:60]}")
            break
    
    # 3. DATOS DE MUESTRA
    if header_row:
        print(f"\n  DATOS DE MUESTRA (filas {header_row+1} a {min(header_row+10, ws.max_row)}):")
        for row_idx in range(header_row + 1, min(header_row + 11, ws.max_row + 1)):
            row_data = {}
            valores = []
            for col_idx in range(1, min(ws.max_column + 1, 25)):
                cell = ws.cell(row=row_idx, column=col_idx)
                col_letter = get_column_letter(col_idx)
                val = cell.value
                
                if val is not None:
                    if isinstance(val, (int, float)):
                        valores.append(f"{col_letter}{row_idx}={val}")
                    else:
                        str_val = str(val)[:50]
                        valores.append(f"{col_letter}{row_idx}='{str_val}'")
                    row_data[col_letter] = {
                        'valor': str(val)[:100] if val else None,
                        'tipo': type(val).__name__
                    }
            
            if valores:
                print(f"    Fila {row_idx}: {', '.join(valores[:15])}")
                analisis['datos_muestra'].append({
                    'fila': row_idx,
                    'datos': row_data
                })
    
    # 4. FÓRMULAS DETALLADAS
    print(f"\nFÓRMULAS:")
    formulas_por_tipo = defaultdict(list)
    formulas_count = 0
    formulas_unicas = set()
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == 'f' and cell.value:
                formulas_count += 1
                formula = str(cell.value)
                formulas_unicas.add(formula)
                
                # Clasificar fórmula
                tipo = 'OTRA'
                if 'VLOOKUP' in formula:
                    tipo = 'VLOOKUP'
                elif 'IF' in formula:
                    tipo = 'IF'
                elif 'SUM' in formula or 'SUMIF' in formula:
                    tipo = 'SUM'
                elif 'MAX' in formula or 'MIN' in formula:
                    tipo = 'MAX/MIN'
                elif 'CONCATENATE' in formula or '&' in formula:
                    tipo = 'CONCATENATE'
                elif 'INDEX' in formula or 'MATCH' in formula:
                    tipo = 'INDEX/MATCH'
                elif 'COUNT' in formula:
                    tipo = 'COUNT'
                elif 'AVERAGE' in formula:
                    tipo = 'AVERAGE'
                elif 'IFERROR' in formula:
                    tipo = 'IFERROR'
                
                formulas_por_tipo[tipo].append({
                    'celda': cell.coordinate,
                    'formula': formula[:300]
                })
                
                # Detectar referencias cruzadas
                if "'" in formula:
                    partes = formula.split("'")
                    for i in range(1, len(partes), 2):
                        if i < len(partes):
                            hoja_ref = partes[i]
                            if hoja_ref and hoja_ref != nombre_hoja:
                                analisis['referencias_cruzadas'].append({
                                    'celda': cell.coordinate,
                                    'referencia': hoja_ref,
                                    'formula': formula[:200]
                                })
    
    print(f"  Total de fórmulas: {formulas_count}")
    print(f"  Fórmulas únicas: {len(formulas_unicas)}")
    print(f"  Distribución por tipo:")
    for tipo, formulas in sorted(formulas_por_tipo.items(), key=lambda x: len(x[1]), reverse=True):
        print(f"    {tipo}: {len(formulas)}")
        if len(formulas) <= 10:
            for f in formulas[:10]:
                print(f"      {f['celda']}: {f['formula'][:80]}")
        else:
            print(f"      (mostrando primeras 5 de {len(formulas)})")
            for f in formulas[:5]:
                print(f"      {f['celda']}: {f['formula'][:80]}")
    
    analisis['formulas'] = {
        'total': formulas_count,
        'unicas': len(formulas_unicas),
        'por_tipo': {k: len(v) for k, v in formulas_por_tipo.items()},
        'ejemplos': {k: v[:10] for k, v in formulas_por_tipo.items()}
    }
    
    # 5. VALIDACIONES
    if hasattr(ws, 'data_validations'):
        validations = list(ws.data_validations.dataValidation)
        if validations:
            print(f"\nVALIDACIONES DE DATOS: {len(validations)}")
            for val in validations:
                val_info = {
                    'rango': str(val.ranges),
                    'tipo': val.type,
                    'formula1': val.formula1 if hasattr(val, 'formula1') else None,
                    'formula2': val.formula2 if hasattr(val, 'formula2') else None,
                    'showErrorMessage': val.showErrorMessage if hasattr(val, 'showErrorMessage') else None,
                    'errorTitle': val.errorTitle if hasattr(val, 'errorTitle') else None
                }
                analisis['validaciones'].append(val_info)
                print(f"  Rango: {val.ranges}")
                print(f"    Tipo: {val.type}")
                if val.formula1:
                    print(f"    Fórmula: {val.formula1[:100]}")
    
    # 6. FORMATOS CONDICIONALES
    if hasattr(ws, 'conditional_formatting'):
        cf_rules = ws.conditional_formatting._cf_rules
        if cf_rules:
            print(f"\nFORMATOS CONDICIONALES: {len(cf_rules)} rangos")
            total_reglas = sum(len(rules) for rules in cf_rules.values())
            print(f"  Total de reglas: {total_reglas}")
            
            for range_str, rules in list(cf_rules.items())[:10]:
                cf_info = {
                    'rango': str(range_str),
                    'num_reglas': len(rules),
                    'reglas': []
                }
                for rule in rules[:3]:  # Primeras 3 reglas
                    rule_info = {
                        'tipo': rule.type if hasattr(rule, 'type') else 'unknown',
                        'formula': rule.formula if hasattr(rule, 'formula') else None
                    }
                    cf_info['reglas'].append(rule_info)
                analisis['formatos_condicionales'].append(cf_info)
                print(f"  Rango: {range_str}, Reglas: {len(rules)}")
    
    # 7. ANÁLISIS DE COLUMNAS CON DATOS
    print(f"\nANÁLISIS DE COLUMNAS:")
    columnas_con_datos = {}
    for col_idx in range(1, min(ws.max_column + 1, 50)):
        col_letter = get_column_letter(col_idx)
        valores_col = []
        tipos_col = defaultdict(int)
        
        for row_idx in range(1, min(ws.max_row + 1, 100)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                valores_col.append(cell.value)
                tipos_col[type(cell.value).__name__] += 1
        
        if valores_col:
            no_vacios = len(valores_col)
            col_info = {
                'columna': col_letter,
                'valores_no_vacios': no_vacios,
                'tipos': dict(tipos_col),
                'ejemplos': [str(v)[:50] for v in valores_col[:5]]
            }
            columnas_con_datos[col_letter] = col_info
            
            if no_vacios > 5:
                print(f"  Columna {col_letter}: {no_vacios} valores, tipos: {dict(tipos_col)}")
                print(f"    Ejemplos: {col_info['ejemplos']}")
    
    analisis['columnas'] = columnas_con_datos
    
    # 8. RANGOS NOMBRADOS (si existen)
    print(f"\nRANGOS NOMBRADOS:")
    # Esto requiere acceso al workbook completo, se hará después
    
    return analisis

def main():
    archivo = r"Herramienta de gestión de riesgo Talento Humano V1 revisada.xlsm"
    
    if not os.path.exists(archivo):
        print(f"ERROR: No se encuentra el archivo: {archivo}")
        return
    
    print("="*120)
    print("ANÁLISIS COMPLETO HOJA POR HOJA")
    print("="*120)
    
    wb = load_workbook(archivo, data_only=False, keep_vba=True)
    
    todas_las_hojas = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        analisis = analizar_hoja_completa(ws, sheet_name)
        todas_las_hojas[sheet_name] = analisis
    
    # Resumen general
    print("\n" + "="*120)
    print("RESUMEN GENERAL")
    print("="*120)
    print(f"Total de hojas analizadas: {len(wb.sheetnames)}")
    
    total_formulas = sum(h['formulas']['total'] for h in todas_las_hojas.values())
    total_validaciones = sum(len(h['validaciones']) for h in todas_las_hojas.values())
    total_cf = sum(len(h['formatos_condicionales']) for h in todas_las_hojas.values())
    total_referencias = sum(len(h['referencias_cruzadas']) for h in todas_las_hojas.values())
    
    print(f"Total de fórmulas: {total_formulas}")
    print(f"Total de validaciones: {total_validaciones}")
    print(f"Total de formatos condicionales: {total_cf}")
    print(f"Total de referencias cruzadas: {total_referencias}")
    
    print("\nHojas ordenadas por complejidad (número de fórmulas):")
    hojas_ordenadas = sorted(
        todas_las_hojas.items(),
        key=lambda x: x[1]['formulas']['total'],
        reverse=True
    )
    
    for nombre, info in hojas_ordenadas:
        if info['formulas']['total'] > 0:
            print(f"  {nombre}: {info['formulas']['total']} fórmulas")
    
    # Guardar análisis completo
    try:
        # Limpiar objetos no serializables
        for hoja_nombre, hoja_info in todas_las_hojas.items():
            # Convertir formatos condicionales a formato serializable
            if 'formatos_condicionales' in hoja_info:
                hoja_info['formatos_condicionales'] = [
                    {
                        'rango': str(cf.get('rango', '')),
                        'num_reglas': cf.get('num_reglas', 0)
                    }
                    for cf in hoja_info['formatos_condicionales']
                ]
        
        with open('analisis_completo_hojas.json', 'w', encoding='utf-8') as f:
            json.dump({
                'total_hojas': len(wb.sheetnames),
                'hojas': todas_las_hojas,
                'resumen': {
                    'total_formulas': total_formulas,
                    'total_validaciones': total_validaciones,
                    'total_formatos_condicionales': total_cf,
                    'total_referencias_cruzadas': total_referencias
                }
            }, f, indent=2, ensure_ascii=False, default=str)
        
        print(f"\nAnálisis guardado en: analisis_completo_hojas.json")
    except Exception as e:
        print(f"\nError al guardar JSON: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()

