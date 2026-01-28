"""
Extracción detallada de todas las fórmulas y lógica de cada hoja
"""
import os
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from collections import defaultdict

def analizar_formulas_detalladas(ws, nombre_hoja):
    """Extrae todas las fórmulas con contexto completo"""
    
    analisis = {
        'nombre': nombre_hoja,
        'formulas_completas': [],
        'formulas_por_celda': {},
        'dependencias': [],
        'referencias_externas': [],
        'logica_negocio': []
    }
    
    print(f"\n{'='*120}")
    print(f"ANÁLISIS DETALLADO DE FÓRMULAS: {nombre_hoja}")
    print(f"{'='*120}")
    
    formulas_encontradas = []
    referencias = set()
    
    # Analizar cada celda con fórmula
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == 'f' and cell.value:
                formula = str(cell.value)
                coord = cell.coordinate
                
                # Obtener valor calculado si existe
                try:
                    valor_calculado = ws[coord].value
                except:
                    valor_calculado = None
                
                # Analizar tipo de fórmula
                tipo_formula = detectar_tipo_formula(formula)
                
                # Extraer referencias
                refs = extraer_referencias(formula, nombre_hoja)
                for ref in refs:
                    referencias.add(str(ref))  # Convertir a string para poder usar en set
                
                formula_info = {
                    'celda': coord,
                    'formula': formula,
                    'valor_calculado': valor_calculado,
                    'tipo': tipo_formula,
                    'referencias': refs,
                    'fila': cell.row,
                    'columna': cell.column
                }
                
                formulas_encontradas.append(formula_info)
                analisis['formulas_por_celda'][coord] = formula_info
                
                # Analizar lógica de negocio
                logica = analizar_logica_negocio(formula, tipo_formula)
                if logica:
                    analisis['logica_negocio'].append({
                        'celda': coord,
                        'formula': formula,
                        'logica': logica
                    })
    
    analisis['formulas_completas'] = formulas_encontradas
    analisis['total_formulas'] = len(formulas_encontradas)
    
    # Analizar dependencias
    for formula_info in formulas_encontradas:
        for ref in formula_info['referencias']:
            if ref['tipo'] == 'celda_misma_hoja':
                analisis['dependencias'].append({
                    'desde': formula_info['celda'],
                    'hacia': ref['referencia'],
                    'formula': formula_info['formula']
                })
            elif ref['tipo'] == 'hoja_externa':
                analisis['referencias_externas'].append({
                    'celda': formula_info['celda'],
                    'hoja_origen': nombre_hoja,
                    'hoja_destino': ref['hoja'],
                    'referencia': ref['referencia'],
                    'formula': formula_info['formula']
                })
    
    # Mostrar resultados
    print(f"\nTotal de fórmulas encontradas: {len(formulas_encontradas)}")
    
    if formulas_encontradas:
        print(f"\nFÓRMULAS DETALLADAS:")
        print(f"{'-'*120}")
        
        # Agrupar por tipo
        por_tipo = defaultdict(list)
        for f in formulas_encontradas:
            por_tipo[f['tipo']].append(f)
        
        for tipo, formulas in sorted(por_tipo.items(), key=lambda x: len(x[1]), reverse=True):
            print(f"\n{tipo.upper()} ({len(formulas)} fórmulas):")
            for f in formulas[:20]:  # Mostrar primeras 20 de cada tipo
                print(f"  {f['celda']}: {f['formula'][:100]}")
                if f['referencias']:
                    print(f"    → Referencias: {', '.join([r['referencia'] for r in f['referencias'][:5]])}")
                if f['valor_calculado'] is not None:
                    print(f"    → Valor: {f['valor_calculado']}")
        
        # Mostrar referencias externas
        if analisis['referencias_externas']:
            print(f"\n{'='*120}")
            print(f"REFERENCIAS EXTERNAS ({len(analisis['referencias_externas'])}):")
            print(f"{'-'*120}")
            refs_por_hoja = defaultdict(list)
            for ref in analisis['referencias_externas']:
                refs_por_hoja[ref['hoja_destino']].append(ref)
            
            for hoja_dest, refs in refs_por_hoja.items():
                print(f"\n  → {hoja_dest} ({len(refs)} referencias):")
                for ref in refs[:10]:
                    print(f"    {ref['celda']} referencia a {ref['referencia']}")
    
    return analisis

def detectar_tipo_formula(formula):
    """Detecta el tipo de fórmula"""
    formula_upper = formula.upper()
    
    if 'VLOOKUP' in formula_upper:
        return 'VLOOKUP'
    elif 'IFERROR' in formula_upper:
        return 'IFERROR'
    elif 'IF(' in formula_upper or 'IF ' in formula_upper:
        if 'AND(' in formula_upper or 'OR(' in formula_upper:
            return 'IF_COMPLEJO'
        return 'IF'
    elif 'SUM' in formula_upper:
        return 'SUM'
    elif 'MAX' in formula_upper:
        return 'MAX'
    elif 'MIN' in formula_upper:
        return 'MIN'
    elif 'CONCATENATE' in formula_upper or '&' in formula:
        return 'CONCATENATE'
    elif 'INDEX' in formula_upper or 'MATCH' in formula_upper:
        return 'INDEX_MATCH'
    elif 'COUNT' in formula_upper:
        return 'COUNT'
    elif 'AVERAGE' in formula_upper:
        return 'AVERAGE'
    elif "'" in formula and '!' in formula:
        return 'REFERENCIA_HOJA'
    elif any(op in formula for op in ['+', '-', '*', '/', '^']):
        return 'CALCULO_MATEMATICO'
    else:
        return 'OTRA'

def extraer_referencias(formula, hoja_actual):
    """Extrae todas las referencias de una fórmula"""
    referencias = []
    
    # Referencias a otras hojas
    if "'" in formula:
        partes = formula.split("'")
        for i in range(1, len(partes), 2):
            if i < len(partes):
                hoja_ref = partes[i]
                if '!' in partes[i+1] if i+1 < len(partes) else False:
                    ref_celda = partes[i+1].split('!')[1].split()[0] if i+1 < len(partes) else ''
                    referencias.append({
                        'tipo': 'hoja_externa',
                        'hoja': hoja_ref,
                        'referencia': ref_celda,
                        'completa': f"'{hoja_ref}'!{ref_celda}"
                    })
    
    # Referencias a celdas en la misma hoja (patrón A1, $A$1, etc.)
    import re
    patron_celda = r'\$?[A-Z]+\$?\d+'
    celdas = re.findall(patron_celda, formula)
    for celda in celdas:
        if celda not in [r['referencia'] for r in referencias]:
            referencias.append({
                'tipo': 'celda_misma_hoja',
                'referencia': celda,
                'hoja': hoja_actual
            })
    
    # Referencias a rangos nombrados (si existen)
    if 'Listas!' in formula or 'Formulas!' in formula:
        # Extraer nombre de rango
        partes = formula.split('!')
        if len(partes) > 1:
            rango = partes[1].split(',')[0].split(')')[0]
            referencias.append({
                'tipo': 'rango_nombrado',
                'hoja': partes[0].replace("'", ""),
                'referencia': rango
            })
    
    return referencias

def analizar_logica_negocio(formula, tipo):
    """Analiza la lógica de negocio de una fórmula"""
    logica = {}
    
    if tipo == 'IF_COMPLEJO':
        # Extraer condiciones
        if 'AND(' in formula:
            logica['tipo'] = 'condicion_multiple_AND'
        elif 'OR(' in formula:
            logica['tipo'] = 'condicion_multiple_OR'
        
        # Casos especiales conocidos
        if 'AD11=2,W11=2' in formula or 'AD=2,W=2' in formula:
            logica['caso_especial'] = 'Si impacto=2 y probabilidad=2, resultado=3.99'
            logica['implementacion'] = 'if (impacto === 2 && probabilidad === 2) return 3.99; else return impacto * probabilidad;'
        
        if 'Riesgo con consecuencia positiva' in formula:
            logica['caso_especial'] = 'Riesgo positivo siempre es nivel bajo'
            logica['implementacion'] = 'if (clasificacion === "Riesgo con consecuencia positiva") return "NIVEL BAJO";'
    
    elif tipo == 'VLOOKUP':
        # Extraer tabla de búsqueda
        if 'Listas!' in formula:
            logica['tabla'] = 'Listas'
            logica['tipo_busqueda'] = 'busqueda_en_catalogo'
        elif 'Formulas!' in formula:
            logica['tabla'] = 'Formulas'
            logica['tipo_busqueda'] = 'busqueda_en_parametros'
        elif 'Mapa de riesgos' in formula:
            logica['tabla'] = 'Mapa de riesgos'
            logica['tipo_busqueda'] = 'busqueda_en_mapa'
    
    elif tipo == 'CONCATENATE':
        logica['tipo'] = 'generacion_id'
        logica['implementacion'] = 'Concatenar campos para generar ID único'
    
    elif tipo == 'MAX':
        if 'AE11:AE20' in formula or 'AO11:AO20' in formula:
            logica['tipo'] = 'maximo_riesgo'
            logica['implementacion'] = 'Encontrar el máximo riesgo en un rango'
    
    return logica if logica else None

def main():
    archivo = r"Herramienta de gestión de riesgo Talento Humano V1 revisada.xlsm"
    
    if not os.path.exists(archivo):
        print(f"ERROR: No se encuentra el archivo: {archivo}")
        return
    
    print("="*120)
    print("EXTRACCIÓN DETALLADA DE FÓRMULAS Y LÓGICA DE NEGOCIO")
    print("="*120)
    
    wb = load_workbook(archivo, data_only=False, keep_vba=True)
    
    todas_las_formulas = {}
    
    # Analizar solo hojas con fórmulas
    hojas_con_formulas = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        tiene_formulas = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f' and cell.value:
                    tiene_formulas = True
                    break
            if tiene_formulas:
                break
        if tiene_formulas:
            hojas_con_formulas.append(sheet_name)
    
    print(f"\nHojas con fórmulas: {len(hojas_con_formulas)}")
    print(f"Hojas: {', '.join(hojas_con_formulas)}")
    
    for sheet_name in hojas_con_formulas:
        ws = wb[sheet_name]
        analisis = analizar_formulas_detalladas(ws, sheet_name)
        todas_las_formulas[sheet_name] = analisis
    
    # Guardar análisis completo
    try:
        # Limpiar para JSON
        for hoja, info in todas_las_formulas.items():
            # Convertir sets a listas
            if 'referencias' in info:
                for f in info.get('formulas_completas', []):
                    if 'referencias' in f:
                        f['referencias'] = [dict(r) for r in f['referencias']]
        
        with open('formulas_detalladas.json', 'w', encoding='utf-8') as f:
            json.dump(todas_las_formulas, f, indent=2, ensure_ascii=False, default=str)
        
        print(f"\n{'='*120}")
        print("Análisis guardado en: formulas_detalladas.json")
    except Exception as e:
        print(f"\nError al guardar: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()

